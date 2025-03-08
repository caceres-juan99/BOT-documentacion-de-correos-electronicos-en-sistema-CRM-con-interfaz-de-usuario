import pyautogui  # Automatización GUI
import shutil  # Manipulación de archivos y directorios
import asyncio  # Corutinas y tareas asincrónicas
import sys  # Funcionalidad del intérprete de Python
import os  # Interacción con el sistema operativo
import time  # Funciones relacionadas con el tiempo
import openpyxl  # Manejo de datos en Excel
import logging

from StorageFunctions import (OpenCRM, ConnectToCRM, CloseCRM, showDesktop, make_noise, get_username_os, terminateProcess, getCurrentDateAndTime)
from ProcessOTP import (searchUpdateOTPControl)
from datetime import datetime
#from dotenv import load_dotenv

async def main():
     
    #load_dotenv() 

    # CONSTANTES
    INPUT_DIRECTORY = "C:/Users/jcacd/OneDrive - Nae/BPO/BOT BOTONES DE PROYECTO/Insumos"
    INPUT_FILENAME = 'Input BOT 001 V1.xlsx'
    GENERIC_ERROR_MSG = 'No se ha procesado el incidente - Para ver más detalles ver el archivo'

    # VARIABLES        
    counter = 1
    resp_funcion = None
    total_rows = 0
    total_cols = 0

    try:
        # Inicio del proceso
        print("=============================== BOT ACTUALIZACION ESTÁNDAR ================================")                        
        print("Iniciando proceso de cargue. Al finalizar, sonará un pitido tres veces.")
        actividad_rpa_selected = "Actualización CRM Estándar"

        # Abrir archivo Excel
        file = os.path.join(INPUT_DIRECTORY, INPUT_FILENAME)
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Eliminar hoja de respaldo anterior y crear una nueva
        if 'Input Backup' in wb.sheetnames:
            del wb['Input Backup']
        
        target_ws = wb.copy_worksheet(ws)
        target_ws.title = "Input Backup"

        total_rows = len(ws['A'])
        total_cols = len(ws[1])
        print("Total de registros en base fuente: ", total_rows - 1)
        print("Total de columnas en base fuente: ", total_cols)

        # Abrir y conectar al sistema CRM
        OpenCRM()
        time.sleep(1)
        ConnectToCRM()       

        # Procesamiento de datos en el bucle
        rows = ws.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=total_cols)

        for row in rows:  # CORREGIDO: Este bloque estaba mal indentado
            if len(row) < total_cols:
                print(f"Fila incompleta detectada en la fila {counter + 1}: {row}")
                continue

            # Validamos el campo 'Status' antes de procesar
            if row[2].value and (row[2].value == 'COMPLETADO' or str(row[2].value).find("ERROR") != -1):
                continue

            # Procesamos solo si la fila es válida
            print('Validando registro: ', str(row[0].value), str(row[1].value), row[3].value, row[4].value, str(row[5].value), row[7].value)
            resp_funcion = searchUpdateOTPControl(
                str(row[0].value), 
                str(row[1].value), 
                row[3].value, 
                row[4].value, 
                str(row[5].value), 
                row[7].value
            )
            print("Respuesta de la función:", resp_funcion)
        
            # Según el valor de resp_funcion:
                                                #0: Se marca como 'COMPLETADO'.
                                                #7: Se marca como 'ERROR'.
                                                #Otro valor: Se marca como 'ERROR - ' seguido de un mensaje genérico (GENERIC_ERROR_MSG).
                                            
            if resp_funcion == 0:
                row[2].value = 'COMPLETADO'
            elif resp_funcion == 7:
                row[2].value = 'ERROR'
            else:
                row[2].value = 'ERROR - ' + GENERIC_ERROR_MSG
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                time.sleep(2)
                OpenCRM()
                ConnectToCRM()
        
            # Log y control del contador
            print("Registro (counter) #: ", counter)
            if counter % 10 == 0:
                print(f'Se han procesado {counter} registros. Reiniciando CRM.')
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                time.sleep(2)
                OpenCRM()
                ConnectToCRM()
        
            counter += 1
            wb.save(file)  # Guarda el archivo después de procesar cada fila
        
        print("Fin del proceso macro")

    except FileNotFoundError as e:
        print(f"Error: No se encontró el archivo {INPUT_FILENAME}. Detalles: {e}")
        exit_program()

    except Exception as e:
        print(f"Error inesperado: {e}")
        exit_program()

    else:
        # Cerrar CRM y finalizar el proceso
        CloseCRM()
        time.sleep(1)
        showDesktop()   
        print("¡Proceso ejecutado exitosamente!")
        print(f"FIN - RPA '{actividad_rpa_selected}'. Usuario [{get_username_os()}] Hora de finalización: {getCurrentDateAndTime()}")

    finally: 
        time.sleep(2)
        make_noise() 
        wb.save(file)  # Asegurar que el archivo se guarde al final

def exit_program():
    sys.exit(0)    

if __name__ == "__main__":
    asyncio.run(main())